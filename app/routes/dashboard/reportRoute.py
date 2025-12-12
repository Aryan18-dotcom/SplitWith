from flask import Blueprint, render_template, request, flash, redirect, url_for, send_file, request
from datetime import datetime
from ...models.expenseModel import ExpenseModel
from ...models.groupModel import GroupModel
from ...models.userModel import UserModel
from ..userAuth import get_session_user
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart

report_bp = Blueprint("reports", __name__, template_folder="templates")


# --------------------------------------------------------
# REPORTS HOME PAGE
# --------------------------------------------------------
@report_bp.route("/reports")
def reports_home():
    user_session = get_session_user()
    if not user_session:
        return render_template(
            "user_auth/login.html",
            message="Please login first.",
            category="error"
        )

    user_id = str(user_session["user_id"])

    try:
        current_user = UserModel.get_user_by_ID(user_id)
    except Exception:
        flash("Unable to load user data.", "error")
        return redirect(url_for("user_auth.login"))

    # Get all user's expenses to determine year range
    user_expenses = ExpenseModel.get_expenses_for_user(user_id)
    if user_expenses:
        years = [e["created_at"].year for e in user_expenses if e.get("created_at")]
        min_year = min(years)
        max_year = max(years)
    else:
        min_year = max_year = datetime.now().year

    groups = GroupModel.get_user_groups(user_id)

    context = {
        "current_user": current_user,
        "current_user_id": user_id,
        "groups": groups,
        "min_year": min_year,
        "max_year": max_year,
        "current_month": datetime.now().month,
        "current_year": datetime.now().year
    }

    return render_template("dashboard/reports.html", **context)



# --------------------------------------------------------
# EXPENSE SUMMARY REPORT
# --------------------------------------------------------
@report_bp.route("/reports/summary")
def report_summary():
    user_id = request.args.get("user_id")

    if not user_id:
        return "User not found", 400

    user_id = str(user_id)
    groups = GroupModel.get_user_groups(user_id)

    total_expense = 0
    total_paid = 0
    total_owe = 0
    total_owed = 0

    # Loop through each group
    for g in groups:

        group_id = str(g["_id"])
        expenses = ExpenseModel.get_expenses_for_group(group_id)

        for e in expenses:

            amount = float(e.get("amount", 0))
            total_expense += amount

            fs = e.get("final_split", {}) or {}
            user_data = fs.get(user_id, {"paid": 0, "net_balance": 0})

            paid = float(user_data.get("paid", 0))
            net_balance = float(user_data.get("net_balance", 0))

            total_paid += paid

            # Negative means YOU NEED TO PAY
            if net_balance < 0:
                total_owe += abs(net_balance)

            # Positive means YOU SHOULD RECEIVE
            elif net_balance > 0:
                total_owed += net_balance

    return {
        "total_expenses": round(total_expense, 2),
        "you_paid": round(total_paid, 2),
        "you_owe": round(total_owe, 2),
        "you_are_owed": round(total_owed, 2)
    }


# --------------------------------------------------------
# MONTHLY REPORT
# --------------------------------------------------------
@report_bp.route("/reports/monthly")
def report_monthly():
    user_id = request.args.get("user_id", "").strip()
    month = int(request.args.get("month"))
    year = int(request.args.get("year"))

    start = datetime(year, month, 1)
    end = datetime(year, month + 1, 1) if month < 12 else datetime(year + 1, 1, 1)

    groups = GroupModel.get_user_groups(user_id)
    expenses_list = []

    for g in groups:
        group_id = str(g["_id"])
        expenses = ExpenseModel.get_expenses_for_group(group_id)

        for e in expenses:
            created_at = e.get("created_at")
            if not created_at:
                continue

            if isinstance(created_at, str):
                created_at = datetime.fromisoformat(created_at)

            # Filter by month AND user
            if start <= created_at < end and (user_id in e.get("split_with", []) or user_id == e.get("created_by")):
                expenses_list.append({
                    "group": g["group_title"],
                    "title": e.get("title", "Untitled"),
                    "amount": float(e.get("amount", 0)),
                    "date": created_at.strftime("%d %b %Y")
                })

    return {"monthly_expenses": expenses_list}





# --------------------------------------------------------
# GROUP WISE REPORT
# --------------------------------------------------------
@report_bp.route("/reports/group/<group_id>")
def report_group(group_id):
    user_id = request.args.get("user_id")
    month = int(request.args.get("month"))
    year = int(request.args.get("year"))

    group = GroupModel.get_group_by_id(group_id)
    if not group:
        return "Group not found", 404

    start = datetime(year, month, 1)
    end = datetime(year, month + 1, 1) if month < 12 else datetime(year + 1, 1, 1)

    expenses = ExpenseModel.get_expenses_for_group(group_id)
    formatted = []

    for e in expenses:
        created_at = e.get("created_at")
        if not created_at:
            continue

        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)

        # Filter by month
        if not (start <= created_at < end):
            continue

        # Calculate what the current user owes/should receive
        fs = e.get("final_split", {}) or {}
        user_data = fs.get(user_id, {"should_pay": 0, "paid": 0, "net_balance": 0})
        owes = user_data["should_pay"] - user_data["paid"] if user_data["should_pay"] - user_data["paid"] > 0 else 0
        owed = abs(user_data["should_pay"] - user_data["paid"]) if user_data["should_pay"] - user_data["paid"] < 0 else 0

        formatted.append({
            "title": e.get("title", "Untitled"),
            "description": e.get("description", ""),
            "amount": float(e.get("amount", 0)),
            "created_at": created_at.strftime("%d %b %Y"),
            "created_by": e.get("created_by"),
            "split_with": e.get("split_with", []),
            "custom_payments": e.get("custom_payments", {}),
            "custom_shares": e.get("custom_shares", {}),
            "you_owe": owes,
            "you_are_owed": owed
        })

    return {
        "group_title": group["group_title"],
        "expenses": formatted
    }

# -------------------------
# HELPER FUNCTION TO CREATE EXCEL
# -------------------------
def create_excel(expenses_data, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = [
        "Group", "Title", "Amount", "You Owe", "You Are Owed",
        "Date", "Created By", "Description", "Split With"
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="800000")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Fill data
    for idx, e in enumerate(expenses_data, start=2):
        ws.cell(row=idx, column=1, value=e.get("group"))
        ws.cell(row=idx, column=2, value=e.get("title"))
        ws.cell(row=idx, column=3, value=e.get("amount", 0))
        ws.cell(row=idx, column=4, value=e.get("you_owe", 0))
        ws.cell(row=idx, column=5, value=e.get("you_are_owed", 0))
        ws.cell(row=idx, column=6, value=e.get("date"))
        ws.cell(row=idx, column=7, value=e.get("created_by"))
        ws.cell(row=idx, column=8, value=e.get("description", ""))
        ws.cell(row=idx, column=9, value=", ".join(e.get("split_with", [])))

        for col in range(1, 10):
            cell = ws.cell(row=idx, column=col)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border

    # Auto-adjust column width
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 5, 50)

    return wb


# -------------------------
# MONTHLY EXCEL
# -------------------------
@report_bp.route("/reports/excel/month")
def monthly_excel_report():
    user_id = request.args.get("user_id")
    month = int(request.args.get("month"))
    year = int(request.args.get("year"))

    if not user_id:
        return "User ID required", 400

    start = datetime(year, month, 1)
    end = datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)

    expenses_data = []
    groups = GroupModel.get_user_groups(user_id)

    for g in groups:
        expenses = ExpenseModel.get_expenses_for_group(g["_id"])

        for e in expenses:
            created_at = e.get("created_at")
            if not created_at:
                continue

            if isinstance(created_at, str):
                try:
                    created_at = datetime.fromisoformat(created_at)
                except Exception:
                    continue

            # Only include expenses in range AND relevant to user
            if not (start <= created_at < end):
                continue
            if not (user_id in e.get("split_with", []) or user_id == e.get("created_by")):
                continue

            fs = e.get("final_split", {}) or {}
            user_data = fs.get(user_id, {"should_pay": 0, "paid": 0})
            owes = max(user_data.get("should_pay", 0) - user_data.get("paid", 0), 0)
            owed = max(user_data.get("paid", 0) - user_data.get("should_pay", 0), 0)

            expenses_data.append({
                "group": g["group_title"],
                "title": e.get("title", "Untitled"),
                "amount": e.get("amount", 0),
                "you_owe": owes,
                "you_are_owed": owed,
                "date": created_at.strftime("%d-%b-%Y"),
                "created_by": e.get("created_by"),
                "description": e.get("description", ""),
                "split_with": e.get("split_with", [])
            })

    wb = create_excel(expenses_data, sheet_name=f"{month}-{year} Expenses")

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    filename = f"Expense_Report_{month}_{year}.xlsx"

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# -------------------------
# YEARLY EXCEL
# -------------------------
@report_bp.route("/reports/excel/year")
def yearly_excel_report():
    user_id = request.args.get("user_id")
    year = int(request.args.get("year"))

    if not user_id:
        return "User ID required", 400

    start = datetime(year, 1, 1)
    end = datetime(year + 1, 1, 1)

    expenses_data = []
    groups = GroupModel.get_user_groups(user_id)

    for g in groups:
        expenses = ExpenseModel.get_expenses_for_group(g["_id"])

        for e in expenses:
            created_at = e.get("created_at")
            if not created_at:
                continue

            if isinstance(created_at, str):
                try:
                    created_at = datetime.fromisoformat(created_at)
                except Exception:
                    continue

            if not (start <= created_at < end):
                continue
            if not (user_id in e.get("split_with", []) or user_id == e.get("created_by")):
                continue

            fs = e.get("final_split", {}) or {}
            user_data = fs.get(user_id, {"should_pay": 0, "paid": 0})
            owes = max(user_data.get("should_pay", 0) - user_data.get("paid", 0), 0)
            owed = max(user_data.get("paid", 0) - user_data.get("should_pay", 0), 0)

            expenses_data.append({
                "group": g["group_title"],
                "title": e.get("title", "Untitled"),
                "amount": e.get("amount", 0),
                "you_owe": owes,
                "you_are_owed": owed,
                "date": created_at.strftime("%d-%b-%Y"),
                "created_by": e.get("created_by"),
                "description": e.get("description", ""),
                "split_with": e.get("split_with", [])
            })

    wb = create_excel(expenses_data, sheet_name=f"{year} Expenses")

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    filename = f"Expense_Report_{year}.xlsx"

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
